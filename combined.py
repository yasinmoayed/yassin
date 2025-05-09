# --- Begin activator.py ---
import hashlib
import json
from cryptography.fernet import Fernet
from tkinter import Tk, Button, filedialog
from datetime import datetime, timedelta
import os
import sys  # اضافه کردن ماژول sys برای خروج کامل
import tkinter as tk
from tkinter import ttk
from bidi.algorithm import get_display
import arabic_reshaper
import sqlite3
import openpyxl
from openpyxl.styles import Alignment
from tkinter import messagebox, simpledialog
# مسیر دایرکتوری فعلی
CURRENT_DIRECTORY = os.path.dirname(os.path.abspath(__file__))
ACTIVATION_STATUS_FILE = os.path.join(CURRENT_DIRECTORY, "activation_status.json")

# Define a valid Base64 key for Fernet encryption
SECRET_KEY = b'Y2hvb3NlQVN0cm9uZ0JhU2U2NEVuY3J5cHRpb25LZXk='  # Replace with your valid Fernet key
cipher = Fernet(SECRET_KEY)
# پیدا کردن مسیر دایرکتوری فایل اجرایی یا کد
if getattr(sys, 'frozen', False):  # اگر برنامه به EXE تبدیل شده باشد
    CURRENT_DIRECTORY = os.path.dirname(sys.executable)
else:  # در حالت اجرای مستقیم کد پایتون
    CURRENT_DIRECTORY = os.path.dirname(os.path.abspath(__file__))

# مسیر فایل activation_status.json
ACTIVATION_STATUS_FILE = os.path.join(CURRENT_DIRECTORY, "activation_status.json")

def save_activation_status(status):
    """ذخیره وضعیت فعال‌سازی در فایل JSON."""
    with open(ACTIVATION_STATUS_FILE, "w") as file:
        json.dump(status, file)
    print(f"Activation status saved to: {ACTIVATION_STATUS_FILE}")

def load_activation_status():
    """بارگذاری وضعیت فعال‌سازی از فایل JSON."""
    if not os.path.exists(ACTIVATION_STATUS_FILE):  # اگر فایل وجود نداشت
        print("Activation file not found.")
        return {"activated": False}
    with open(ACTIVATION_STATUS_FILE, "r") as file:
        return json.load(file)

def check_existing_activation():
    """بررسی فعال‌سازی قبلی."""
    status = load_activation_status()
    if status.get("activated"):
        print("برنامه قبلاً فعال شده است.")
        return True
    return False
def load_encrypted_file(file_path):
    """
    Load and decrypt the encrypted file.
    """
    with open(file_path, "rb") as file:
        encrypted_data = file.read()
    decrypted_data = cipher.decrypt(encrypted_data).decode()
    return json.loads(decrypted_data)

def save_activation_status(expiration_date):  # noqa: F811
    """
    Save the activation status and expiration date to a file.
    """
    status_data = {
        "activated": True,
        "expiration_date": expiration_date.strftime("%Y-%m-%d")
    }
    with open(ACTIVATION_STATUS_FILE, "w") as file:
        json.dump(status_data, file)

def load_activation_status():  # noqa: F811
    """
    Load the activation status from file.
    """
    if not os.path.exists(ACTIVATION_STATUS_FILE):
        return {"activated": False}

    with open(ACTIVATION_STATUS_FILE, "r") as file:
        return json.load(file)

def validate_activation_data(data):
    """
    Validate the activation data.
    """
    data_string = json.dumps({
        "name": data["name"],
        "email": data["email"],
        "duration_days": data["duration_days"],
        "issue_date": data["issue_date"],
    }, separators=(",", ":")) + "MySecretKey"
    expected_code = hashlib.sha256(data_string.encode()).hexdigest()[:16]

    if data["activation_code"] != expected_code:
        return False, "Invalid activation code!"

    # Check expiration date
    issue_date = datetime.strptime(data["issue_date"], "%Y-%m-%d")
    expiration_date = issue_date + timedelta(days=data["duration_days"])
    if datetime.now() > expiration_date:
        return False, "Activation code has expired!"

    return True, expiration_date

def on_activate():
    file_path = filedialog.askopenfilename(filetypes=[("Key Files", "*.key")])
    if not file_path:
        messagebox.showerror("Error", "No file selected!")
        sys.exit()  # خروج کامل از برنامه اگر هیچ فایلی انتخاب نشده باشد

    try:
        data = load_encrypted_file(file_path)
    except Exception as e:
        messagebox.showerror("Error", f"Failed to load file: {str(e)}")
        sys.exit()  # خروج کامل از برنامه اگر فایل خراب باشد

    valid, result = validate_activation_data(data)
    if valid:
        save_activation_status(result)  # Save expiration date
        messagebox.showinfo("Success", "Activation successful!")
        root.destroy()  # Close the window
        run_main_program()  # اجرای برنامه اصلی در صورت موفقیت فعال‌سازی
    else:
        messagebox.showerror("Error", result)
        sys.exit()  # خروج کامل از برنامه اگر کد فعال‌سازی نامعتبر باشد
def run_main_program():
    """اجرای برنامه اصلی."""
    print("برنامه اصلی با موفقیت اجرا شد!")
def check_existing_activation():  # noqa: F811
    """
    Check if activation is already valid.
    """
    status = load_activation_status()
    if status.get("activated"):
        expiration_date = datetime.strptime(status["expiration_date"], "%Y-%m-%d")
        if datetime.now() <= expiration_date:
            messagebox.showinfo("Info", "Program already activated.")
            root.destroy()
            run_main_program()  # اجرای مستقیم برنامه اصلی اگر قبلاً فعال شده باشد
            return True
    return False

# Build the GUI
root = Tk()
root.title("Activator")

# Check if already activated
if check_existing_activation():
    pass  # برنامه اصلی قبلاً اجرا شده است
else:
    activate_button = Button(root, text="Activate", command=on_activate)
    activate_button.pack(pady=20)

    # اضافه کردن متد برای بستن برنامه در صورت بسته شدن پنجره اکتیو
    def on_close():
        sys.exit()  # خروج کامل از برنامه در صورت بسته شدن پنجره اکتیو

    root.protocol("WM_DELETE_WINDOW", on_close)  # هندل بسته شدن پنجره
    root.mainloop()


# --- End main.py ---
# =======================================================================================================================================================================================
#region پیشنیاز ها
# ======================================================================================================================================================================================

# مسیر فایل لایسنس
CURRENT_DIRECTORY = os.path.dirname(os.path.abspath(__file__))
LICENSE_FILE = os.path.join(CURRENT_DIRECTORY, "activation_status.json")

def load_license_status():
    """وضعیت لایسنس را بارگذاری می‌کند."""
    if not os.path.exists(LICENSE_FILE):
        return {"activated": False}

    with open(LICENSE_FILE, "r") as file:
        return json.load(file)

def save_license_status(status):
    """ذخیره وضعیت لایسنس."""
    with open(LICENSE_FILE, "w") as file:
        json.dump(status, file)

def activate_license():
    """فعال‌سازی لایسنس."""
    activation_code = simpledialog.askstring("فعال‌سازی", "لطفاً کد فعال‌سازی را وارد کنید:")

    if activation_code == "1234-5678-ABCD":  # کد نمونه
        expiration_date = (datetime.now() + timedelta(days=30)).strftime("%Y-%m-%d")
        license_data = {
            "activated": True,
            "expiration_date": expiration_date
        }
        save_license_status(license_data)
        messagebox.showinfo("موفقیت", "لایسنس با موفقیت فعال شد!")
    else:
        messagebox.showerror("خطا", "کد فعال‌سازی اشتباه است.")

def main_program():
    """برنامه اصلی."""
    status = load_license_status()

    if not status.get("activated"):
        activate_license()
        return

    expiration_date = datetime.strptime(status["expiration_date"], "%Y-%m-%d")
    if expiration_date < datetime.now():
        messagebox.showwarning("هشدار", "لایسنس منقضی شده است!")
        activate_license()
        return

    messagebox.showinfo("خوش آمدید", "برنامه با موفقیت اجرا شد!")
# =======================================================================================================================================================================================
#region اضافه کردن قابلیت نمایش و حذف لایسنس
# =======================================================================================================================================================================================

ACTIVATION_STATUS_FILE = "activation_status.json"  # فایل وضعیت لایسنس

# مسیر فایل JSON
CURRENT_DIRECTORY = os.path.dirname(os.path.abspath(__file__))
ACTIVATION_STATUS_FILE = os.path.join(CURRENT_DIRECTORY, "activation_status.json")

def load_license_status():  # noqa: F811
    """وضعیت لایسنس را از فایل بارگذاری می‌کند."""
    if not os.path.exists(ACTIVATION_STATUS_FILE):
        return {"activated": False}

    with open(ACTIVATION_STATUS_FILE, "r") as file:
        return json.load(file)

def show_license_info():
    """نمایش اطلاعات لایسنس."""
    status = load_license_status()
    if not status.get("activated"):
        messagebox.showwarning("هشدار", "هیچ لایسنسی فعال نیست.")
        return

    expiration_date = datetime.strptime(status["expiration_date"], "%Y-%m-%d")
    remaining_days = (expiration_date - datetime.now()).days

    license_info = (
        f"نام: {status.get('name', 'نامشخص')}\n"
        f"تاریخ انقضا: {status['expiration_date']}\n"
        f"روزهای باقی‌مانده: {remaining_days} روز"
    )
    messagebox.showinfo("اطلاعات لایسنس", license_info)

def delete_license_status():
    """حذف وضعیت لایسنس."""
    if os.path.exists(ACTIVATION_STATUS_FILE):
        os.remove(ACTIVATION_STATUS_FILE)
        messagebox.showinfo("اطلاع", "لایسنس با موفقیت حذف شد.")
    else:
        messagebox.showwarning("هشدار", "هیچ لایسنسی برای حذف وجود ندارد.")

def show_license_info():  # noqa: F811
    """نمایش اطلاعات لایسنس."""
    status = load_license_status()
    if not status.get("activated"):
        messagebox.showwarning("هشدار", "هیچ لایسنسی فعال نیست.")
        return

    expiration_date = datetime.strptime(status["expiration_date"], "%Y-%m-%d")
    remaining_days = (expiration_date - datetime.now()).days

    license_info = (
        f"نام: {status.get('name', 'نامشخص')}\n"
        f"تاریخ انقضا: {status['expiration_date']}\n"
        f"روزهای باقی‌مانده: {remaining_days} روز"
    )
    messagebox.showinfo("اطلاعات لایسنس", license_info)
#endregion
# =======================================================================================================================================================================================
#region # تنظیمات و متغیرهای ثابت
# فایل ذخیره مواد اولیه پیش‌فرض
DEFAULT_MATERIALS_FILE = "materials_data.json"


# فایل پایگاه داده SQLite
DB_FILE = "materials_data.db"

#endregion
# =======================================================================================================================================================================================
# regionتابع اصلاح نمایش متن فارسی
# =======================================================================================================================================================================================
def reshape_text(text):
    reshaped_text = get_display(arabic_reshaper.reshape(text))
    return reshaped_text
#endregion
# =======================================================================================================================================================================================
# regionتنظیمات پایگاه داده SQLite
# ===========================================================================================================================================================================
def init_db():
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    # ایجاد جدول مواد اولیه
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS materials (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT UNIQUE NOT NULL,
        data TEXT NOT NULL
    )
    """)
    # ایجاد جدول گونه‌ها
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS species (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT UNIQUE NOT NULL,
        data TEXT NOT NULL
    )
    """)
    conn.commit()
    conn.close()

# ذخیره گونه‌ها در پایگاه داده
def save_species_to_db(name, species_data):
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("""
    INSERT OR REPLACE INTO species (name, data)
    VALUES (?, ?)
    """, (name, json.dumps(species_data, ensure_ascii=False)))
    conn.commit()
    conn.close()

# حذف گونه از پایگاه داده
def delete_species_from_db(name):
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("DELETE FROM species WHERE name = ?", (name,))
    conn.commit()
    conn.close()

# بارگذاری گونه‌ها از پایگاه داده
def load_species_from_db():
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("SELECT name, data FROM species")
    rows = cursor.fetchall()
    conn.close()
    return {name: json.loads(data) for name, data in rows}

# ذخیره مواد اولیه در پایگاه داده
def save_material_to_db(name, material_data):
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("""
    INSERT OR REPLACE INTO materials (name, data)
    VALUES (?, ?)
    """, (name, json.dumps(material_data, ensure_ascii=False)))
    conn.commit()
    conn.close()

# حذف مواد اولیه از پایگاه داده
def delete_material_from_db(name):
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("DELETE FROM materials WHERE name = ?", (name,))
    conn.commit()
    conn.close()

# بارگذاری مواد اولیه از پایگاه داده
def load_materials_from_db():
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("SELECT name, data FROM materials")
    rows = cursor.fetchall()
    conn.close()
    return {name: json.loads(data) for name, data in rows}
#



# پر کردن داده‌های مواد اولیه پیش‌فرض در پایگاه داده
def populate_default_materials():

    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    for name, data in default_materials_data.items():
        cursor.execute("""
        INSERT OR IGNORE INTO materials (name, data)
        VALUES (?, ?)
        """, (name, json.dumps(data, ensure_ascii=False)))
    conn.commit()
    conn.close()
# پر کردن داده‌های مواد اولیه پیش‌فرض در پایگاه داده
def populate_default_species():
    
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    for name, data in default_materials_data.items():
        cursor.execute("""
        INSERT OR IGNORE INTO materials (name, data)
        VALUES (?, ?)
        """, (name, json.dumps(data, ensure_ascii=False)))
    conn.commit()
    conn.close()
#endregion
# =======================================================================================================================================================================================
# regionداده‌های پیش‌فرض مواد اولیه
# =======================================================================================================================================================================================
default_materials_data = {
    "کنجاله سویا": {
        "پروتئین خام (%)": 48,
        "لیپیدهای خام (%)": 1.5,
        "فیبر خام (%)": 3.5,
        "انرژی ناخالص (kc/k)": 3100,
        "آرژنین (%)": 3.2,
        "لیزین (%)": 2.7,
        "متیونین (%)": 0.7,
        "کلسیم (%)": 0.3,
        "فسفر (%)": 0.65,
        "گوگرد (%)": 0.4,
        "آهن (mg/kg)": 100,
        "روی (mg/kg)": 60,
        "B3 (mg/kg)": 30,
        "کولین (mg/kg)": 2000,
        "اینوزیتول (mg/kg)": 1000
    },
    "پودر ماهی": {
        "پروتئین خام (%)": 65,
        "لیپیدهای خام (%)": 8,
        "خاکستر (%)": 15,
        "انرژی ناخالص (kc/k)": 4500,
        "آرژنین (%)": 4.5,
        "لیزین (%)": 5.5,
        "متیونین (%)": 2.2,
        "کلسیم (%)": 5.5,
        "فسفر (%)": 3.0,
        "گوگرد (%)": 0.3,
        "آهن (mg/kg)": 120,
        "روی (mg/kg)": 80,
        "B3 (mg/kg)": 20,
        "کولین (mg/kg)": 1000,
        "اینوزیتول (mg/kg)": 500
    },
    "آرد گندم": {
        "پروتئین خام (%)": 11,
        "لیپیدهای خام (%)": 1.5,
        "فیبر خام (%)": 2.5,
        "نشاسته (%)": 58.85,
        "انرژی ناخالص (kc/k)": 3000,
        "لیزین (%)": 0.3,
        "متیونین (%)": 0.2,
        "کلسیم (%)": 0.02,
        "فسفر (%)": 0.12,
        "آهن (mg/kg)": 4.32,
        "B3 (mg/kg)": 3.5,
        "B5 (mg/kg)": 0.8,
        "E (mg/kg)": 0.5
    },
    "پودر گوشت 50 درصد": {
        "پروتئین خام (%)": 50.2,
        "لیپیدهای خام (%)": 10.01,
        "خاکستر (%)": 30.51,
        "انرژی ناخالص (kc/k)": 3881.45,
        "آرژنین (%)": 3.37,
        "لیزین (%)": 3.484,
        "متیونین (%)": 0.98,
        "کلسیم (%)": 9.598,
        "فسفر (%)": 4.839,
        "گوگرد (%)": 0.34,
        "آهن (mg/kg)": 438,
        "روی (mg/kg)": 93,
        "B3 (mg/kg)": 54.2,
        "کولین (mg/kg)": 1998,
        "اینوزیتول (mg/kg)": 100
    },
    "گلوتن ذرت 40 درصد": {
        "پروتئین خام (%)": 40,
        "لیپیدهای خام (%)": 2.5,
        "فیبر خام (%)": 3.5,
        "نشاسته (%)": 11.5,
        "انرژی ناخالص (kc/k)": 3600,
        "لیزین (%)": 0.6,
        "متیونین (%)": 0.5,
        "کلسیم (%)": 0.01,
        "فسفر (%)": 0.15,
        "آهن (mg/kg)": 7.7,
        "B3 (mg/kg)": 1.2,
        "E (mg/kg)": 5
    },
    "گلوتن ذرت 50 درصد": {
        "پروتئین خام (%)": 51.3,
        "لیپیدهای خام (%)": 7.8,
        "نشاسته (%)": 13.1,
        "انرژی ناخالص (kc/k)": 4844.65,
        "آرژنین (%)": 1.54,
        "لیزین (%)": 0.87,
        "متیونین (%)": 1.23,
        "کلسیم (%)": 0.06,
        "فسفر (%)": 0.42,
        "گوگرد (%)": 0.5,
        "آهن (mg/kg)": 332,
        "روی (mg/kg)": 49,
        "B3 (mg/kg)": 49.8,
        "کولین (mg/kg)": 360,
        "اینوزیتول (mg/kg)": 270.22
    },
    "گلوتن گندم 78 درصد": {
        "پروتئین خام (%)": 78.17,
        "لیپیدهای خام (%)": 2.4,
        "نشاسته (%)": 1.5,
        "انرژی ناخالص (kc/k)": 5014.34,
        "آرژنین (%)": 3.09,
        "لیزین (%)": 2.16,
        "متیونین (%)": 1.35,
        "کلسیم (%)": 0.1,
        "فسفر (%)": 0.24,
        "آهن (mg/kg)": 45.5,
        "روی (mg/kg)": 50,
        "B3 (mg/kg)": 100,
        "کولین (mg/kg)": 1500,
        "اینوزیتول (mg/kg)": 142.68
    },
    "پودر خون": {
        "پروتئین خام (%)": 80.0,
        "لیپیدهای خام (%)": 1.5,
        "خاکستر (%)": 5.0,
        "انرژی ناخالص (kc/k)": 3600,
        "آرژنین (%)": 4.0,
        "لیزین (%)": 8.0,
        "متیونین (%)": 0.5,
        "کلسیم (%)": 0.02,
        "فسفر (%)": 0.3,
        "آهن (mg/kg)": 1500.0,
        "روی (mg/kg)": 20.0,
        "B9 (mg/kg)": 1.0,
        "B3 (mg/kg)": 2.0,
        "B6 (mg/kg)": 1.0
    },
    "پودر پر": {
        "پروتئین خام (%)": 80.0,
        "لیپیدهای خام (%)": 2.5,
        "خاکستر (%)": 4.0,
        "انرژی ناخالص (kc/k)": 3100,
        "آرژنین (%)": 3.5,
        "لیزین (%)": 2.0,
        "متیونین (%)": 0.7,
        "کلسیم (%)": 0.3,
        "فسفر (%)": 0.5,
        "گوگرد (%)": 0.3,
        "آهن (mg/kg)": 150.0,
        "منگنز (mg/kg)": 50.0,
        "سلنیوم (mg/kg)": 0.3,
        "روی (mg/kg)": 60.0
    },
    "ضایعات جوجه کشی": {
        "پروتئین خام (%)": 40.0,
        "لیپیدهای خام (%)": 15.0,
        "فیبر خام (%)": 2.0,
        "خاکستر (%)": 20.0,
        "انرژی ناخالص (kc/k)": 2500,
        "آرژنین (%)": 1.8,
        "لیزین (%)": 2.0,
        "متیونین (%)": 0.6,
        "کلسیم (%)": 6.0,
        "فسفر (%)": 3.5,
        "گوگرد (%)": 0.2,
        "آهن (mg/kg)": 180.0,
        "منگنز (mg/kg)": 40.0,
        "سلنیوم (mg/kg)": 0.2,
        "روی (mg/kg)": 50.0
    },
    "زرده تخم مرغ": {
        "پروتئین خام (%)": 16.0,
        "لیپیدهای خام (%)": 32.0,
        "خاکستر (%)": 2.0,
        "انرژی ناخالص (kc/k)": 3700,
        "آرژنین (%)": 1.2,
        "لیزین (%)": 0.9,
        "متیونین (%)": 0.4,
        "کلسیم (%)": 0.1,
        "فسفر (%)": 0.7,
        "گوگرد (%)": 0.2,
        "آهن (mg/kg)": 7.0,
        "سلنیوم (mg/kg)": 0.3,
        "روی (mg/kg)": 5.0,
        "ید (mg/kg)": 0.02,
        "B7 (mg/kg)": 0.05,
        "B9 (mg/kg)": 0.2,
        "B5 (mg/kg)": 0.4,
        "B2 (mg/kg)": 0.3,
        "B1 (mg/kg)": 0.2,
        "A (IU/kg)": 3000,
        "D (IU/kg)": 50,
        "E (mg/kg)": 10.0,
        "کولین (mg/kg)": 2500.0
    },
    "روغن سویا": {
        "لیپیدهای خام (%)": 99,
        "انرژی ناخالص (kc/k)": 9228.01,
        "ید (mg/kg)": 0.05,
        "E (mg/kg)": 61.04,
        "K (mg/kg)": 0.18,
        "کولین (mg/kg)": 2
    },
    "لیسیتین": {
        "لیپیدهای خام (%)": 60,
        "خاکستر (%)": 0.2,
        "انرژی ناخالص (kc/k)": 8000,
        "کلسیم (%)": 0.01,
        "فسفر (%)": 0.02,
        "E (mg/kg)": 10
    },
    "روغن ماهی": {
        "لیپیدهای خام (%)": 99.9,
        "انرژی ناخالص (kc/k)": 9000
    },
    "دی کلسیوم فسفات": {
        "خاکستر (%)": 99,
        "کلسیم (%)": 29.393,
        "فسفر (%)": 22.8
    },
    "مخلوط ویتامینه میگو": {
        "پروتئین خام (%)": 12.6,
        "لیپیدهای خام (%)": 2.4,
        "فیبر خام (%)": 5.6,
        "نشاسته (%)": 25.25,
        "انرژی ناخالص (kc/k)": 4189.77,
        "کلسیم (%)": 0.088,
        "فسفر (%)": 1.048,
        "B7 (mg/kg)": 750.0,
        "B9 (mg/kg)": 2500.0,
        "B3 (mg/kg)": 50000.0,
        "B5 (mg/kg)": 25000.0,
        "B6 (mg/kg)": 20000.0,
        "B2 (mg/kg)": 12500.0,
        "B1 (mg/kg)": 15000.0,
        "B12 (mg/kg)": 4.0,
        "A (IU/kg)": 2000000.0,
        "D (IU/kg)": 1000000.0,
        "E (mg/kg)": 50000.0,
        "K (mg/kg)": 20000.0
    },
    "مخلوط معدنی میگو": {
        "پروتئین خام (%)": 2.0,
        "لیپیدهای خام (%)": 0.32,
        "فیبر خام (%)": 0.2,
        "خاکستر (%)": 75.0,
        "نشاسته (%)": 12.0,
        "انرژی ناخالص (kc/k)": 860.42,
        "کلسیم (%)": 0.015,
        "فسفر (%)": 0.015,
        "مس (mg/kg)": 20000.0,
        "آهن (mg/kg)": 40000.0,
        "منگنز (mg/kg)": 2000.0,
        "سلنیوم (mg/kg)": 280.0,
        "روی (mg/kg)": 40000.0,
        "ید (mg/kg)": 1500.0
    },
    "ژلاتین": {
        "پروتئین خام (%)": 85.0,
        "خاکستر (%)": 2.0,
        "انرژی ناخالص (kc/k)": 3500,
        "آرژنین (%)": 8.0,
        "لیزین (%)": 0.02,
        "متیونین (%)": 0.01,
        "کلسیم (%)": 0.1,
        "فسفر (%)": 0.02,
        "سدیم (%)": 0.5,
        "آهن (mg/kg)": 10.0
    },
    "Vitamin C Phosphate 35%": {
        "C (mg/kg)": 350000.0
    },
    "جوش شیرین": {
        "سدیم (%)": 27.3
    },
    "اسید فایر": {
    },
    "پریبیوتیک": {
        "پروتئین خام (%)": 55.0,
        "لیپیدهای خام (%)": 2.0,
        "مانان‌الیگوساکارید (MOS) (%)": 20.0,
        "بتا گلوکان (%)": 25.0,
        "لیزین (%)": 1.0,
        "متیونین (%)": 0.5,
        "کلسیم (%)": 0.3,
        "فسفر (%)": 0.6,
        "مس (mg/kg)": 15.0,
        "آهن (mg/kg)": 50.0,
        "منگنز (mg/kg)": 30.0,
        "سلنیوم (mg/kg)": 0.3,
        "روی (mg/kg)": 60.0
    },
    "کنسانتره ویتامینه ماهی سی باس": {
        "سلنیوم (mg/kg)": 0.5,
        "C (mg/kg)": 500.0,
        "B7 (mg/kg)": 10.0,
        "B9 (mg/kg)": 30.0,
        "B3 (mg/kg)": 200.0,
        "B5 (mg/kg)": 50.0,
        "B6 (mg/kg)": 20.0,
        "B2 (mg/kg)": 40.0,
        "B1 (mg/kg)": 30.0,
        "B12 (mg/kg)": 0.1,
        "A (IU/kg)": 500000,
        "D (IU/kg)": 100000,
        "E (mg/kg)": 300.0,
        "K (mg/kg)": 100.0,
        "کولین (mg/kg)": 10000.0,
        "اینوزیتول (mg/kg)": 2000.0
    },
    "کنسانتره معدنی سی باس": {
        "خاکستر (%)": 50.0,
        "کلسیم (%)": 24.0,
        "فسفر (%)": 18.0,
        "سدیم (%)": 2.0,
        "پتاسیم (%)": 0.5,
        "منیزیم (%)": 3.0,
        "مس (mg/kg)": 50.0,
        "آهن (mg/kg)": 200.0,
        "منگنز (mg/kg)": 100.0,
        "سلنیوم (mg/kg)": 0.5,
        "روی (mg/kg)": 150.0,
        "ید (mg/kg)": 1.0,
        "کبالت (mg/kg)": 0.5
    },
    "اینوزیتول": {
        "اینوزیتول (mg/kg)": 980000.0
    },
    "بتائین": {
    },
    "ال-آرژنین": {
        "پروتئین خام (%)": 98,
        "انرژی ناخالص (kc/k)": 5528.2,
        "آرژنین (%)": 98
    },
    "ال-هیستیدین": {
        "پروتئین خام (%)": 98,
        "انرژی ناخالص (kc/k)": 5528.2,
        "هیستیدین (%)": 98
    },
    "ال-لیزین": {
        "پروتئین خام (%)": 99,
        "انرژی ناخالص (kc/k)": 5583.17,
        "لیزین (%)": 82.616
    },
    "متیونین": {
        "انرژی ناخالص (kc/k)": 4000,
        "متیونین (%)": 99.0,
        "گوگرد (%)": 21.0
    },
    "ال والین": {
        "پروتئین خام (%)": 72,
        "انرژی ناخالص (kc/k)": 5117.11,
        "والین (%)": 96.005
    },
    "ال-ترئونین": {
        "پروتئین خام (%)": 70.9,
        "انرژی ناخالص (kc/k)": 5100.38,
        "ترئونین (%)": 95.995
    },
    "ال-تریپتوفان": {
        "پروتئین خام (%)": 82.5,
        "انرژی ناخالص (kc/k)": 5284.42,
        "تریپتوفان (%)": 95.997
    }
}
species_data = {
    "Seabass 60-100g (Grower)": {
        "پروتئین خام (%)": 46,
        "لیپیدهای خام (%)": 13,
        "فیبر خام (%)": 3,
        "خاکستر (%)": 10,
        "انرژی ناخالص (kc/k)": 3793,
        "نشاسته (%)": 14,
        "آرژنین (%)": 2.5,
        "لیزین (%)": 3.5,
        "متیونین (%)": 0.9,
        "کلسیم (%)": 1.0,
        "فسفر (%)": 0.8,
        "سدیم (%)": 0.5,
        "کلر (%)": 0.4,
        "پتاسیم (%)": 0.8,
        "منیزیم (%)": 0.2,
        "مس (mg/kg)": 10,
        "گوگرد (%)": 0.6,
        "آهن (mg/kg)": 120,
        "منگنز (mg/kg)": 15,
        "سلنیوم (mg/kg)": 0.2,
        "روی (mg/kg)": 80,
        "ید (mg/kg)": 0.3,
        "کبالت (mg/kg)": 0.1,
        "C (mg/kg)": 250,
        "B7 (mg/kg)": 0.05,
        "B9 (mg/kg)": 1.0,
        "B3 (mg/kg)": 20,
        "B5 (mg/kg)": 8,
        "B6 (mg/kg)": 2,
        "B2 (mg/kg)": 2,
        "B1 (mg/kg)": 1.5,
        "B12 (mg/kg)": 0.01,
        "A (IU/kg)": 8000,
        "D (IU/kg)": 2500,
        "E (mg/kg)": 70,
        "K (mg/kg)": 1.5,
        "کولین (mg/kg)": 2200,
        "اینوزیتول (mg/kg)": 1200
    },
    "Seabass 200-400g (Grower)": {
    "پروتئین خام (%)": 45,
    "لیپیدهای خام (%)": 13,
    "فیبر خام (%)": 7,
    "خاکستر (%)": 9,
    "نشاسته (%)": 14,
    "انرژی ناخالص (kc/k)": 3820,
    "آرژنین (%)": 1.95,
    "هیستیدین (%)": 0.57,
    "ایزولوسین (%)": 1.52,
    "لوسین (%)": 2.67,
    "لیزین (%)": 2.38,
    "متیونین (%)": 0.91,
    "فنیل آلانین (%)": 1.29,
    "ترئونین (%)": 1.30,
    "تریپتوفان (%)": 0.37,
    "والین (%)": 1.56,
    "جمع n-3 (%)": 0.56,
    "جمع n-6 (%)": 0.9,
    "فسفولیپیدها (%)": 0.7,
    "کلسیم (%)": 1.0,
    "فسفر (%)": 1.00,
    "سدیم (%)": 0.08,
    "کلر (%)": 0.08,
    "پتاسیم (%)": 0.2,
    "منیزیم (%)": 0.06,
    "مس (mg/kg)": 6.9,
    "آهن (mg/kg)": 80,
    "منگنز (mg/kg)": 9.8,
    "سلنیوم (mg/kg)": 0.15,
    "روی (mg/kg)": 41,
    "ید (mg/kg)": 2.07,
    "C (mg/kg)": 100,
    "B7 (mg/kg)": 0.20,
    "B9 (mg/kg)": 1.0,
    "B3 (mg/kg)": 26,
    "B5 (mg/kg)": 15,
    "B6 (mg/kg)": 10,
    "B2 (mg/kg)": 6.0,
    "B1 (mg/kg)": 2.0,
    "B12 (mg/kg)": 0.01,
    "A (IU/kg)": 0.90,
    "D (IU/kg)": 10,
    "E (mg/kg)": 150,
    "K (mg/kg)": 3.0,
    "کولین (mg/kg)": 1000,
    "اینوزیتول (mg/kg)": 150
    },
    "Seabass 400-700g (Grower)": {
        "پروتئین خام (%)": 44,
        "لیپیدهای خام (%)": 14,
        "فیبر خام (%)": 7,
        "خاکستر (%)": 9,
        "نشاسته (%)": 14,
        "انرژی ناخالص (kc/k)": 3799,
        "آرژنین (%)": 1.86,
        "هیستیدین (%)": 0.55,
        "ایزولوسین (%)": 1.46,
        "لوسین (%)": 2.57,
        "لیزین (%)": 2.28,
        "متیونین (%)": 0.87,
        "فنیل آلانین (%)": 1.24,
        "ترئونین (%)": 1.25,
        "تریپتوفان (%)": 0.36,
        "والین (%)": 1.49,
        "جمع n-3 (%)": 0.51,
        "جمع n-6 (%)": 0.9,
        "فسفولیپیدها (%)": 0.6,
        "کلسیم (%)": 1.0,
        "فسفر (%)": 0.98,
        "سدیم (%)": 0.08,
        "کلر (%)": 0.08,
        "پتاسیم (%)": 0.2,
        "منیزیم (%)": 0.06,
        "مس (mg/kg)": 6.4,
        "آهن (mg/kg)": 75,
        "منگنز (mg/kg)": 9.2,
        "سلنیوم (mg/kg)": 0.14,
        "روی (mg/kg)": 37,
        "ید (mg/kg)": 1.85,
        "C (mg/kg)": 100,
        "B7 (mg/kg)": 0.20,
        "B9 (mg/kg)": 1.0,
        "B3 (mg/kg)": 26,
        "B5 (mg/kg)": 15,
        "B6 (mg/kg)": 10,
        "B2 (mg/kg)": 6.0,
        "B1 (mg/kg)": 2.0,
        "B12 (mg/kg)": 0.01,
        "A (IU/kg)": 0.90,
        "D (IU/kg)": 10,
        "E (mg/kg)": 150,
        "K (mg/kg)": 3.0,
        "کولین (mg/kg)": 1000,
        "اینوزیتول (mg/kg)": 150
    },
    "Seabass 700-1000g (Grower)": {
        "پروتئین خام (%)": 43,
        "لیپیدهای خام (%)": 14,
        "فیبر خام (%)": 7,
        "خاکستر (%)": 9,
        "نشاسته (%)": 14,
        "انرژی ناخالص (kc/k)": 3817,
        "آرژنین (%)": 1.83,
        "هیستیدین (%)": 0.55,
        "ایزولوسین (%)": 1.44,
        "لوسین (%)": 2.53,
        "لیزین (%)": 2.24,
        "متیونین (%)": 0.86,
        "فنیل آلانین (%)": 1.22,
        "ترئونین (%)": 1.23,
        "تریپتوفان (%)": 0.35,
        "والین (%)": 1.47,
        "جمع n-3 (%)": 0.40,
        "جمع n-6 (%)": 0.9,
        "فسفولیپیدها (%)": 0.5,
        "کلسیم (%)": 2.0,
        "فسفر (%)": 0.98,
        "سدیم (%)": 0.07,
        "کلر (%)": 0.07,
        "پتاسیم (%)": 0.2,
        "منیزیم (%)": 0.06,
        "مس (mg/kg)": 6.2,
        "آهن (mg/kg)": 74,
        "منگنز (mg/kg)": 9.0,
        "سلنیوم (mg/kg)": 0.13,
        "روی (mg/kg)": 34,
        "ید (mg/kg)": 1.71,
        "C (mg/kg)": 100,
        "B7 (mg/kg)": 0.20,
        "B9 (mg/kg)": 1.0,
        "B3 (mg/kg)": 26,
        "B5 (mg/kg)": 15,
        "B6 (mg/kg)": 10,
        "B2 (mg/kg)": 6.0,
        "B1 (mg/kg)": 2.0,
        "B12 (mg/kg)": 0.01,
        "A (IU/kg)": 0.90,
        "D (IU/kg)": 10,
        "E (mg/kg)": 150,
        "K (mg/kg)": 3.0,
        "کولین (mg/kg)": 1000,
        "اینوزیتول (mg/kg)": 150
    },
    "Whiteleg Shrimp1-3g (Starter)": {
        "رطوبت (%)": 10,
        "پروتئین خام (%)": 38,
        "لیپیدهای خام (%)": 6,
        "فیبر خام (%)": 5,
        "خاکستر (%)": 9,
        "نشاسته (%)": 16,
        "انرژی ناخالص (kc/k)": 3389,
        "آرژنین (%)": 2.52,
        "هیستیدین (%)": 0.70,
        "ایزولوسین (%)": 1.28,
        "لوسین (%)": 2.21,
        "لیزین (%)": 2.63,
        "متیونین (%)": 0.78,
        "فنیل آلانین (%)": 1.52,
        "ترئونین (%)": 1.26,
        "تریپتوفان (%)": 0.33,
        "والین (%)": 1.60,
        "جمع n-3 (%)": 0.41,
        "جمع n-6 (%)": 0.5,
        "فسفولیپیدها (%)": 1.1,
        "کلسیم (%)": 2,
        "فسفر (%)": 1.01,
        "سدیم (%)": 0.09,
        "کلر (%)": 0.09,
        "پتاسیم (%)": 0.2,
        "منیزیم (%)": 0.08,
        "مس (mg/kg)": 8.1,
        "آهن (mg/kg)": 106,
        "منگنز (mg/kg)": 11.8,
        "سلنیوم (mg/kg)": 0.27,
        "روی (mg/kg)": 48,
        "ید (mg/kg)": 3.19,
        "C (mg/kg)": 150,
        "B7 (mg/kg)": 0.50,
        "B9 (mg/kg)": 5.0,
        "B3 (mg/kg)": 26,
        "B5 (mg/kg)": 15,
        "B6 (mg/kg)": 10,
        "B2 (mg/kg)": 15.0,
        "B1 (mg/kg)": 4.0,
        "B12 (mg/kg)": 0.01,
        "A (IU/kg)": 1.40,
        "D (IU/kg)": 30,
        "E (mg/kg)": 150,
        "K (mg/kg)": 3.0,
        "کولین (mg/kg)": 1000,
        "اینوزیتول (mg/kg)": 150
    },
    "Whiteleg Shrimp 3-6g (Pre-grower)": {
        "رطوبت (%)": 10,
        "پروتئین خام (%)": 37,
        "لیپیدهای خام (%)": 6,
        "فیبر خام (%)": 5,
        "خاکستر (%)": 8,  # مقدار نامشخص
        "نشاسته (%)": 16,
        "انرژی ناخالص (kc/k)": 3365,
        "آرژنین (%)": 2.49,
        "هیستیدین (%)": 0.68,
        "ایزولوسین (%)": 1.24,
        "لوسین (%)": 2.06,
        "لیزین (%)": 2.53,
        "متیونین (%)": 0.71,
        "فنیل آلانین (%)": 1.42,
        "ترئونین (%)": 1.22,
        "تریپتوفان (%)": 0.33,
        "والین (%)": 1.50,
        "جمع n-3 (%)": 0.30,
        "جمع n-6 (%)": 0.4,
        "فسفولیپیدها (%)": 0.9,
        "کلسیم (%)": 1,  # مقدار نامشخص
        "فسفر (%)": 0.95,
        "سدیم (%)": 0.08,
        "کلر (%)": 0.08,
        "پتاسیم (%)": 0.2,
        "منیزیم (%)": 0.07,
        "مس (mg/kg)": 7.6,
        "آهن (mg/kg)": 98,
        "منگنز (mg/kg)": 11.0,
        "سلنیوم (mg/kg)": 0.25,
        "روی (mg/kg)": 45,
        "ید (mg/kg)": 2.93,
        "C (mg/kg)": 150,
        "B7 (mg/kg)": 0.50,
        "B9 (mg/kg)": 5.0,
        "B3 (mg/kg)": 26,
        "B5 (mg/kg)": 15,
        "B6 (mg/kg)": 10,
        "B2 (mg/kg)": 15.0,
        "B1 (mg/kg)": 4.0,
        "B12 (mg/kg)": 0.01,
        "A (IU/kg)": 0.90,
        "D (IU/kg)": 10,
        "E (mg/kg)": 150,
        "K (mg/kg)": 3.0,
        "کولین (mg/kg)": 1000,
        "اینوزیتول (mg/kg)": 150
    },
    "Whiteleg Shrimp 6-12g (Grower)": {
        "رطوبت (%)": 10,
        "پروتئین خام (%)": 37,
        "لیپیدهای خام (%)": 6,
        "فیبر خام (%)": 5,
        "خاکستر (%)": 8,  # مقدار نامشخص
        "نشاسته (%)": 16,
        "انرژی ناخالص (kc/k)": 3348,
        "آرژنین (%)": 2.48,
        "هیستیدین (%)": 0.67,
        "ایزولوسین (%)": 1.21,
        "لوسین (%)": 1.96,
        "لیزین (%)": 2.46,
        "متیونین (%)": 0.67,
        "فنیل آلانین (%)": 1.36,
        "ترئونین (%)": 1.19,
        "تریپتوفان (%)": 0.32,
        "والین (%)": 1.44,
        "جمع n-3 (%)": 0.25,
        "جمع n-6 (%)": 0.4,
        "فسفولیپیدها (%)": 0.8,
        "کلسیم (%)": 1,  # مقدار نامشخص
        "فسفر (%)": 0.88,
        "سدیم (%)": 0.07,
        "کلر (%)": 0.07,
        "پتاسیم (%)": 0.2,
        "منیزیم (%)": 0.07,
        "مس (mg/kg)": 7.1,
        "آهن (mg/kg)": 90,
        "منگنز (mg/kg)": 10.3,
        "سلنیوم (mg/kg)": 0.22,
        "روی (mg/kg)": 41,
        "ید (mg/kg)": 2.65,
        "C (mg/kg)": 200,
        "B7 (mg/kg)": 0.50,
        "B9 (mg/kg)": 5.0,
        "B3 (mg/kg)": 26,
        "B5 (mg/kg)": 15,
        "B6 (mg/kg)": 10,
        "B2 (mg/kg)": 15.0,
        "B1 (mg/kg)": 4.0,
        "B12 (mg/kg)": 0.01,
        "A (IU/kg)": 0.90,
        "D (IU/kg)": 10,
        "E (mg/kg)": 150,
        "K (mg/kg)": 3.0,
        "کولین (mg/kg)": 1000,
        "اینوزیتول (mg/kg)": 150
    },
    "Whiteleg Shrimp >12g (Finisher)": {
        "رطوبت (%)": 10,
        "پروتئین خام (%)": 36,
        "لیپیدهای خام (%)": 6,
        "فیبر خام (%)": 5,
        "خاکستر (%)": 8,  # مقدار نامشخص
        "نشاسته (%)": 16,
        "انرژی ناخالص (kc/k)": 3305,
        "آرژنین (%)": 2.46,
        "هیستیدین (%)": 0.66,
        "ایزولوسین (%)": 1.18,
        "لوسین (%)": 1.87,
        "لیزین (%)": 2.40,
        "متیونین (%)": 0.63,
        "فنیل آلانین (%)": 1.30,
        "ترئونین (%)": 1.17,
        "تریپتوفان (%)": 0.32,
        "والین (%)": 1.38,
        "جمع n-3 (%)": 0.22,
        "جمع n-6 (%)": 0.4,
        "فسفولیپیدها (%)": 0.8,
        "کلسیم (%)": 1,  # مقدار نامشخص
        "فسفر (%)": 0.75,
        "سدیم (%)": 0.07,
        "کلر (%)": 0.07,
        "پتاسیم (%)": 0.2,
        "منیزیم (%)": 0.06,
        "مس (mg/kg)": 6.8,
        "آهن (mg/kg)": 83,
        "منگنز (mg/kg)": 9.7,
        "سلنیوم (mg/kg)": 0.20,
        "روی (mg/kg)": 38,
        "ید (mg/kg)": 2.39,
        "C (mg/kg)": 200,
        "B7 (mg/kg)": 0.50,
        "B9 (mg/kg)": 5.0,
        "B3 (mg/kg)": 26,
        "B5 (mg/kg)": 15,
        "B6 (mg/kg)": 10,
        "B2 (mg/kg)": 15.0,
        "B1 (mg/kg)": 4.0,
        "B12 (mg/kg)": 0.01,
        "A (IU/kg)": 0.90,
        "D (IU/kg)": 10,
        "E (mg/kg)": 150,
        "K (mg/kg)": 3.0,
        "کولین (mg/kg)": 1000,
        "اینوزیتول (mg/kg)": 150
    },
    "Whiteleg Shrimp >20g (Brood)": {
        "رطوبت (%)": 10,
        "پروتئین خام (%)": 37,
        "لیپیدهای خام (%)": 6,
        "فیبر خام (%)": 5,
        "خاکستر (%)": 8,  # مقدار نامشخص
        "نشاسته (%)": 16,
        "انرژی ناخالص (kc/k)": 3365,
        "آرژنین (%)": 2.49,
        "هیستیدین (%)": 0.68,
        "ایزولوسین (%)": 1.24,
        "لوسین (%)": 2.06,
        "لیزین (%)": 2.53,
        "متیونین (%)": 0.71,
        "فنیل آلانین (%)": 1.42,
        "ترئونین (%)": 1.22,
        "تریپتوفان (%)": 0.33,
        "والین (%)": 1.50,
        "جمع n-3 (%)": 0.19,
        "جمع n-6 (%)": 0.4,
        "فسفولیپیدها (%)": 0.8,
        "کلسیم (%)": 1,  # مقدار نامشخص
        "فسفر (%)": 0.95,
        "سدیم (%)": 0.08,
        "کلر (%)": 0.08,
        "پتاسیم (%)": 0.2,
        "منیزیم (%)": 0.07,
        "مس (mg/kg)": 7.6,
        "آهن (mg/kg)": 98,
        "منگنز (mg/kg)": 11.0,
        "سلنیوم (mg/kg)": 0.25,
        "روی (mg/kg)": 45,
        "ید (mg/kg)": 2.93,
        "C (mg/kg)": 150,
        "B7 (mg/kg)": 0.50,
        "B9 (mg/kg)": 5.0,
        "B3 (mg/kg)": 26,
        "B5 (mg/kg)": 15,
        "B6 (mg/kg)": 10,
        "B2 (mg/kg)": 15.0,
        "B1 (mg/kg)": 4.0,
        "B12 (mg/kg)": 0.01,
        "A (IU/kg)": 0.90,
        "D (IU/kg)": 10,
        "E (mg/kg)": 150,
        "K (mg/kg)": 3.0,
        "کولین (mg/kg)": 1000,
        "اینوزیتول (mg/kg)": 150
    }
}
#endregion
# =======================================================================================================================================================================================
# regionرابط کاربری (GUI)
# ===========================================================================================================================================================================
def run_main_program():  # noqa: F811
    """اجرای برنامه اصلی."""
    messagebox.showinfo("برنامه اصلی", "برنامه اصلی با موفقیت اجرا شد!")

class DietCalculatorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("محاسبه جیره غذایی")

        # ترتیب استاندارد پارامترها
        self.standard_order = [
            "پروتئین خام (%)",
            "لیپیدهای خام (%)",
            "فیبر خام (%)",
            "خاکستر (%)",
            "نشاسته (%)",
            "انرژی ناخالص (kc/k)",
            "آرژنین (%)",
            "لیزین (%)",
            "متیونین (%)",
            "کلسیم (%)",
            "فسفر (%)",
            "سدیم (%)",
            "کلر (%)",
            "پتاسیم (%)",
            "منیزیم (%)",
            "مس (mg/kg)",
            "گوگرد (%)",
            "آهن (mg/kg)",
            "منگنز (mg/kg)",
            "سلنیوم (mg/kg)",
            "روی (mg/kg)",
            "ید (mg/kg)",
            "کبالت (mg/kg)",
            "C (mg/kg)",
            "B7 (mg/kg)",
            "B9 (mg/kg)",
            "B3 (mg/kg)",
            "B5 (mg/kg)",
            "B6 (mg/kg)",
            "B2 (mg/kg)",
            "B1 (mg/kg)",
            "B12 (mg/kg)",
            "A (IU/kg)",
            "D (IU/kg)",
            "E (mg/kg)",
            "K (mg/kg)",
            "کولین (mg/kg)",
            "اینوزیتول (mg/kg)"
        ]

        # بارگذاری گونه‌ها از پایگاه داده
        self.species_data = load_species_from_db()

        # بخش گونه
        tk.Label(root, text=reshape_text("گونه")).grid(row=0, column=0, padx=10, pady=10)
        self.species_combobox = ttk.Combobox(root, values=list(species_data.keys()), width=30)
        self.species_combobox.grid(row=0, column=1, padx=10, pady=10)
        self.species_combobox.set("")

        # دکمه‌های مربوط به گونه
        tk.Button(root, text=reshape_text("اضافه کردن گونه جدید"), command=self.open_add_species_window).grid(row=0, column=2, padx=5, pady=5)
        tk.Button(root, text=reshape_text("مدیریت گونه‌ها"), command=self.manage_species_window).grid(row=0, column=3, padx=5, pady=5)

        # بخش مواد اولیه
        self.materials_frame = tk.Frame(root)
        self.materials_frame.grid(row=1, column=0, columnspan=4, padx=5, pady=5)
        self.materials_widgets = []
        self.add_material_row()
        #جدید
        tk.Button(root, text=reshape_text("نمایش اطلاعات لایسنس"), command=show_license_info).grid(row=5, column=0, padx=5, pady=5)
        tk.Button(root, text=reshape_text("حذف لایسنس"), command=delete_license_status).grid(row=5, column=1, padx=5, pady=5)
        #
        tk.Button(root, text=reshape_text("اضافه کردن ماده اولیه"), command=self.add_material_row).grid(row=2, column=0, padx=5, pady=5)
        tk.Button(root, text=reshape_text("محاسبه جیره"), command=self.calculate_diet).grid(row=2, column=1, padx=5, pady=5)
        tk.Button(root, text=reshape_text("اضافه کردن ماده اولیه جدید"), command=self.open_add_material_window).grid(row=2, column=2, padx=5, pady=5)
        tk.Button(root, text=reshape_text("مدیریت مواد اولیه"), command=self.manage_materials_window).grid(row=2, column=3, padx=5, pady=5)

        # جدول نتایج
        self.results_table = ttk.Treeview(root, columns=("param", "calculated", "standard", "difference"), show="headings", height=15)
        self.results_table.grid(row=3, column=0, columnspan=4, padx=5, pady=5)
        self.results_table.heading("param", text=reshape_text("پارامتر"))
        self.results_table.heading("calculated", text=reshape_text("محاسبه‌شده"))
        self.results_table.heading("standard", text=reshape_text("استاندارد"))
        self.results_table.heading("difference", text=reshape_text("تفاوت"))

        self.results_table.tag_configure("less_than", foreground="red")
        self.results_table.tag_configure("greater_than", foreground="blue")

       # به‌روزرسانی Combobox گونه‌ها
        self.update_species_combobox()

    def update_species_combobox(self):
        self.species_combobox["values"] = list(self.species_data.keys(), width=50)

    def add_material_row(self):
        row = len(self.materials_widgets)
        material_combobox = ttk.Combobox(self.materials_frame, values=list(materials_data.keys()), width=30)
        material_combobox.grid(row=row, column=0, padx=5, pady=5)
        material_combobox.set(list(materials_data.keys())[0] if materials_data else "")

        percentage_entry = tk.Entry(self.materials_frame)
        percentage_entry.grid(row=row, column=1, padx=5, pady=5)
        percentage_entry.insert(0, "0")

        self.materials_widgets.append((material_combobox, percentage_entry))

    def open_add_material_window(self):
        add_window = tk.Toplevel(self.root)
        add_window.title("اضافه کردن ماده اولیه جدید")

        # ایجاد کانواس و اسکرول‌بار
        canvas = tk.Canvas(add_window)
        scrollbar = tk.Scrollbar(add_window, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        tk.Label(scrollable_frame, text=reshape_text("نام ماده اولیه:")).grid(row=0, column=0, padx=5, pady=5)
        material_name_entry = tk.Entry(scrollable_frame, width=80)
        material_name_entry.grid(row=0, column=1, padx=5, pady=5)

        param_entries = {}
        for i, param in enumerate(self.standard_order):
            tk.Label(scrollable_frame, text=reshape_text(param + ":")).grid(row=i+1, column=0, padx=5, pady=2)
            entry = tk.Entry(scrollable_frame, width=80)
            entry.grid(row=i+1, column=1, padx=5, pady=2)
            param_entries[param] = entry

        def save_material():
            material_name = material_name_entry.get()
            if not material_name:
                messagebox.showerror("خطا", "لطفاً نام ماده اولیه را وارد کنید.")
                return

            try:
                new_material = {}
                for param, entry in param_entries.items():
                    value = entry.get()
                    new_material[param] = float(value) if value.strip() else 0.0

                save_material_to_db(material_name, new_material)
                materials_data[material_name] = new_material
                messagebox.showinfo("موفقیت", f"ماده '{material_name}' با موفقیت اضافه شد.")

                self.update_material_combobox()
                add_window.destroy()
            except ValueError:
                messagebox.showerror("خطا", "مقادیر وارد شده باید عددی باشند.")

        tk.Button(scrollable_frame, text=reshape_text("ذخیره"), command=save_material).grid(
            row=len(self.standard_order)+1, column=0, columnspan=2, pady=10
        )

    def manage_materials_window(self):
        manage_window = tk.Toplevel(self.root)
        manage_window.title("مدیریت مواد اولیه")

        materials_listbox = tk.Listbox(manage_window, width=80, height=20)
        materials_listbox.pack(side="left", fill="y", padx=10, pady=10)

        scrollbar = tk.Scrollbar(manage_window, orient="vertical", command=materials_listbox.yview)
        scrollbar.pack(side="right", fill="y")

        materials_listbox.config(yscrollcommand=scrollbar.set)

        for material in materials_data.keys():
            materials_listbox.insert("end", material)

        def delete_material():
            selected = materials_listbox.curselection()
            if not selected:
                messagebox.showerror("خطا", "لطفاً یک ماده را انتخاب کنید.")
                return

            material_name = materials_listbox.get(selected)
            delete_material_from_db(material_name)
            del materials_data[material_name]
            materials_listbox.delete(selected)
            self.update_material_combobox()
            messagebox.showinfo("موفقیت", f"ماده '{material_name}' حذف شد.")

        tk.Button(manage_window, text=reshape_text("حذف"), command=delete_material).pack(pady=5)

    def open_add_species_window(self):
        add_window = tk.Toplevel(self.root)
        add_window.title("اضافه کردن گونه جدید")

        # ایجاد کانواس و اسکرول‌بار
        canvas = tk.Canvas(add_window)
        scrollbar = tk.Scrollbar(add_window, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        tk.Label(scrollable_frame, text=reshape_text("نام گونه:")).grid(row=0, column=0, padx=5, pady=5)
        species_name_entry = tk.Entry(scrollable_frame, width=80)
        species_name_entry.grid(row=0, column=1, padx=5, pady=5)

        param_entries = {}
        for i, param in enumerate(self.standard_order):
            tk.Label(scrollable_frame, text=reshape_text(param + ":")).grid(row=i+1, column=0, padx=5, pady=2)
            entry = tk.Entry(scrollable_frame, width=50)
            entry.grid(row=i+1, column=1, padx=5, pady=2)
            param_entries[param] = entry

        def save_species():
            species_name = species_name_entry.get()
            if not species_name:
                messagebox.showerror("خطا", "لطفاً نام گونه را وارد کنید.")
                return

            try:
                new_species = {}
                for param, entry in param_entries.items():
                    value = entry.get()
                    new_species[param] = float(value) if value.strip() else 0.0

                save_species_to_db(species_name, new_species)
                species_data[species_name] = new_species
                messagebox.showinfo("موفقیت", f"گونه '{species_name}' با موفقیت اضافه شد.")

                self.update_species_combobox()
                add_window.destroy()
            except ValueError:
                messagebox.showerror("خطا", "مقادیر وارد شده باید عددی باشند.")

        tk.Button(scrollable_frame, text=reshape_text("ذخیره"), command=save_species).grid(
            row=len(self.standard_order)+1, column=0, columnspan=2, pady=10
        )

    def manage_species_window(self):
        manage_window = tk.Toplevel(self.root)
        manage_window.title("مدیریت گونه‌ها")

        species_listbox = tk.Listbox(manage_window, width=80, height=50)
        species_listbox.pack(side="left", fill="y", padx=10, pady=10)

        scrollbar = tk.Scrollbar(manage_window, orient="vertical", command=species_listbox.yview)
        scrollbar.pack(side="right", fill="y")

        species_listbox.config(yscrollcommand=scrollbar.set)

        for species in species_data.keys():
            species_listbox.insert("end", species)

        def delete_species():
            selected = species_listbox.curselection()
            if not selected:
                messagebox.showerror("خطا", "لطفاً یک گونه را انتخاب کنید.")
                return

            species_name = species_listbox.get(selected)
            delete_species_from_db(species_name)
            del species_data[species_name]
            species_listbox.delete(selected)
            self.update_species_combobox()
            messagebox.showinfo("موفقیت", f"گونه '{species_name}' حذف شد.")

        tk.Button(manage_window, text=reshape_text("حذف"), command=delete_species).pack(pady=5)

    def update_material_combobox(self):
        new_material_list = list(materials_data.keys())
        for combobox, _ in self.materials_widgets:
            combobox["values"] = new_material_list

    def update_species_combobox(self):  # noqa: F811
        self.species_combobox["values"] = list(species_data.keys())

    def calculate_diet(self):
        """محاسبه جیره بر اساس مواد اولیه و درصدهای وارد شده"""
        try:
            selected_species = self.species_combobox.get()
            if selected_species not in species_data:
                raise ValueError("گونه نامعتبر است.")

            # جمع‌آوری اطلاعات مواد اولیه و درصدها
            material_weights = {}
            total_percentage = 0  # جمع درصد کل مواد اولیه
            for combobox, entry in self.materials_widgets:
                material = combobox.get()
                if material not in materials_data:
                    raise ValueError(f"ماده اولیه '{material}' نامعتبر است.")
                try:
                    weight = float(entry.get())
                    if weight < 0 or weight > 100:
                        raise ValueError("درصد وزنی باید بین 0 و 100 باشد.")
                    total_percentage += weight
                    material_weights[material] = weight / 100  # تبدیل به درصد
                except ValueError:
                    raise ValueError("درصد وزنی باید یک عدد معتبر باشد.")

            # بررسی اینکه مجموع درصدها از 100 بیشتر نباشد
            if total_percentage > 100:
                raise ValueError("مجموع درصد مواد اولیه نباید از 100 بیشتر باشد.")

            # محاسبه جیره
            final_composition = {}
            for material, weight in material_weights.items():
                for param, value in materials_data[material].items():
                    if "(%)" in param:
                        # پارامتر درصدی
                        final_composition[param] = final_composition.get(param, 0) + value * weight
                    else:
                        # پارامتر غیر درصدی (تبدیل به واحد مناسب)
                        final_composition[param] = final_composition.get(param, 0) + value * weight * 1

            # مقایسه با استاندارد گونه
            standard_data = species_data[selected_species]
            comparison = {
                param: {
                    "calculated": final_composition.get(param, 0),
                    "standard": standard_data.get(param, 0),
                    "difference": final_composition.get(param, 0) - standard_data.get(param, 0)
                }
                for param in set(final_composition.keys()).union(standard_data.keys())
            }

            # پاک کردن جدول قدیمی
            for item in self.results_table.get_children():
                self.results_table.delete(item)

            # مرتب‌سازی نتایج بر اساس ترتیب استاندارد
            sorted_params = sorted(comparison.keys(), key=lambda x: self.standard_order.index(x) if x in self.standard_order else float('inf'))

            # نمایش نتایج در جدول
            for param in sorted_params:
                values = comparison[param]
                tag = None
                if values["difference"] < 0:
                    tag = "less_than"
                elif values["difference"] > 0:
                    tag = "greater_than"

                self.results_table.insert("", "end", values=(
                    param,
                    round(values["calculated"], 2),
                    round(values["standard"], 2),
                    round(values["difference"], 2)
                ), tags=(tag,))

        except ValueError as e:
            messagebox.showerror("خطا", str(e))
        # دکمه برای ذخیره به اکسل
        tk.Button(root, text=reshape_text("ذخیره به اکسل"), command=self.export_to_excel).grid(row=4, column=0, columnspan=2, pady=10)

    def export_to_excel(self):
        """ذخیره نتایج جدول در فایل اکسل"""
        try:
            # باز کردن پنجره انتخاب مسیر ذخیره‌سازی
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                title="ذخیره فایل اکسل"
            )

            # بررسی اینکه آیا کاربر مسیری انتخاب کرده است
            if not file_path:
                return  # اگر کاربر روی "لغو" کلیک کرده باشد

            # ایجاد فایل اکسل
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Diet Results"

            # افزودن سرستون‌ها
            headers = ["پارامتر", "محاسبه‌شده", "استاندارد", "تفاوت"]
            ws.append(headers)

            # افزودن داده‌ها از جدول
            for item in self.results_table.get_children():
                row = self.results_table.item(item)["values"]
                ws.append(row)

            # تنظیم استایل و تراز متن
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=4):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            # ذخیره فایل در مسیر انتخاب‌شده
            wb.save(file_path)

            # پیام موفقیت
            messagebox.showinfo("ذخیره به اکسل", f"فایل با موفقیت ذخیره شد به آدرس:\n{file_path}")
        except Exception as e:
            messagebox.showerror("خطا", f"خطایی در ذخیره‌سازی پیش آمد: {e}")
#endregion
# =======================================================================================================================================================================================
# regionاجرای اصلی برنامه
# ==============================================================================================================================================================================
def populate_default_materials():  # noqa: F811
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    for name, data in default_materials_data.items():
        cursor.execute("""
        INSERT OR IGNORE INTO materials (name, data)
        VALUES (?, ?)
        """, (name, json.dumps(data, ensure_ascii=False)))
    conn.commit()
    conn.close()


# مقداردهی اولیه پایگاه داده و داده‌های پیش‌فرض
init_db()
populate_default_materials()

# بارگذاری مواد اولیه از پایگاه داده
materials_data = load_materials_from_db()
if __name__ == "__main__":
    root = tk.Tk()
    app = DietCalculatorApp(root)
    root.mainloop()
#endregion
# --- End main.py ---
